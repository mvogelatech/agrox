from openpyxl import load_workbook
import json
import sys

inputFile = sys.argv[1]
outputFile = sys.argv[2]
inputTable = sys.argv[3]

AT = {
    'distanceTitle': 'E7',
    'distance': 'E8'
}

# Read JSON
with open(inputFile, encoding='utf-8') as f:
    modals = json.load(f)

droneCompanies = []
planeCompanies = []
for modal in modals:
    for company in modal['fields'][0]['distances']:
        if modal['modal'] == 'drone':
            droneCompanies.append(company)
        else:
            planeCompanies.append(company)

for modal in modals:
    keep_sheets = ['Apoio']
    if modal['modal'] == 'drone':
        companies = droneCompanies
    else:
        companies = planeCompanies

    for company in companies:

        # Read existing workbook
        prescription_table = load_workbook(filename=inputTable)
        for field in modal['fields']:
            # Grab current worksheet
            current_sheet = prescription_table[field['name']]

            # Add sheet to the keeping list
            keep_sheets.append(field['name'])

            # Add field data
            current_sheet[AT['distanceTitle']] = 'Distância [Km]'
            current_sheet[AT['distance']] = field['distances'][company]

        for sheetName in prescription_table.sheetnames:
            if sheetName not in keep_sheets:
                del prescription_table[sheetName]

        # Save the file
        prescription_table.save(outputFile+ '_' + modal['modal'] + '_' + company + '.xlsx')
