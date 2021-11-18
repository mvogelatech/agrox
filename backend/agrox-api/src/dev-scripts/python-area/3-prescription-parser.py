from openpyxl import load_workbook
import pandas as pd
import json
import sys
import os
import logging
import io

inputFile = sys.argv[1]
outputFolder = sys.argv[2]

# [(droneValues, droneComments), (planeValues, planeComments)]
ranges = [(('H6', 'S15'), ('H17', 'H26')), (('H31', 'S40'), ('H42', 'H51'))]
recommended = 'J2'
author_idx = 'B1'
whatsapp_idx = 'B2'
id = 'B28'

# Read existing workbook
prescription_table = load_workbook(filename=inputFile, data_only=True)
sheets = prescription_table.sheetnames
sheets.remove('Apoio')

for sheet in sheets:
    # Grab current worksheet
    current_sheet = prescription_table[sheet]

    for i, method in enumerate(ranges):
        # Read Values --------------------------------------------------------------------------------------------------
        values_rows = []
        for row in current_sheet[method[0][0]:method[0][1]]:
            values_cols = []
            for cell in row:
                values_cols.append(cell.value)
            values_rows.append(values_cols)

        # Read comments
        comments_rows = []
        for row in current_sheet[method[1][0]:method[1][1]]:
            drone_comments_cols = []
            for cell in row:
                drone_comments_cols.append(cell.value)
            comments_rows.append(drone_comments_cols)

        # Transform into dataframe -------------------------------------------------------------------------------------
        values_df = pd.DataFrame(values_rows)
        comments_df = pd.DataFrame(comments_rows)

        # Filter relevant data -----------------------------------------------------------------------------------------
        values_df.dropna(inplace=True)
        values_df.reset_index(drop=True, inplace=True)
        values_df.drop([1, 4], axis=1, inplace=True)

        comments_df.dropna(inplace=True)
        comments_df.reset_index(drop=True, inplace=True)

        # Rename columns
        values_df.columns = ['plague', 'product', 'dosage', 'syrup', 'buffer', 'nozzle_type', 'drop_size',
                             'pressure', 'total', 'unity']

        comments_df.columns = ['comments']

        comments_list = comments_df['comments'].to_list()

        modal = str(current_sheet[recommended].value).lower()

        if modal == 'drone':
            method_r = '1'
        elif modal == 'avião':
            method_r = '2'
        else:
            method_r = '0'

        diagnosis_id = str(current_sheet[id].value)

        if diagnosis_id == 'None':
             diagnosis_id = '-1'

        # Transform to JSON
        out = '{"products":' + values_df.to_json(orient='records') + ', "comments":' + json.dumps(comments_list, ensure_ascii=False)  + '}'
        if i == 0:
            drone_out = '"drone": ' + out
        elif i == 1:
            plane_out = '"plane": ' + out

        author = current_sheet[author_idx].value
        whatsapp = current_sheet[whatsapp_idx].value

    # Save the file
    with io.open(os.path.join(outputFolder,'field_' + sheet + '.json'), 'w', encoding='utf8') as f:
        f.write(f'{{ {drone_out}, {plane_out}, "recommended_method":{method_r}, "diagnosis_id":{diagnosis_id}, "author":"{author}", "whatsapp":"{whatsapp}"}}')
